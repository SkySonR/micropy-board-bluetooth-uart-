from openpyxl import Workbook
from bluetooth import *
from time import sleep
"""
creating list with valid mac-addresses
define all variables
"""

valid_devices = ["30:14:11:20:02:37", "98:D3:32:30:87:6A"]
active_devices = []
port = 1
size = 1024
wb = Workbook()
ws = wb.active
yes = 0
no = 0
other = 0

# searching for all valid devices

print "performing inquiry..."

nearby_devices = discover_devices(lookup_names = False)

for addr in nearby_devices:
    if addr in valid_devices:
        active_devices.append(addr)
"""
receiving data from pyboardes
counting sumes of different votes
"""
for device in active_devices:
    print device
    sock = BluetoothSocket(RFCOMM)
    sock.connect((device, port))
    data = sock.recv(size)
    if data:
        data = data[0]
        print(data)
        if '+' in data:
            yes += 1
        elif '-' in data:
            no += 1
        else:
            other += 1
# appending all data to xlsx file
        max = ws.max_row
        ws.cell(row=1, column=1, value="Votes")
        ws.cell(row=1, column=2, value="Address")
        for row, entry in enumerate(data.split(),start=1):
            ws.cell(row=row+max, column=1, value=entry)
        for row, entry in enumerate(device.split(),start=1):
            ws.cell(row=row+max, column=2, value=entry)
        wb.save("sample.xlsx")
        sock.send(data)
        sock.close()
        sleep(3)
# define all sumes raws in xlsx file

max = ws.max_row
ws.cell(row=row+max, column=1, value="Sum +")
ws.cell(row=row+max, column=2, value="Sum -")
ws.cell(row=row+max, column=3, value="Sum ~")

# appending all sumes to xlsx file

max +=1
for row, entry in enumerate(str(yes), start=1):
    ws.cell(row=row+max, column=1, value=entry)
for row, entry in enumerate(str(no), start=1):
    ws.cell(row=row+max, column=2, value=entry)
for row, entry in enumerate(str(other), start=1):
    ws.cell(row=row+max, column=3, value=entry)
wb.save("sample.xlsx")
